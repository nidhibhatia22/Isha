import os

import openpyxl
from openpyxl.chart import Reference, BarChart3D
import prettytable as PrettyTable
from Code import config as cfg
import numpy as np
import matplotlib.pyplot as plt



class Generate_tables_and_graphs:

    # accuracy()

    def accuracy(self):

        wb = openpyxl.Workbook()
        ws = wb.active
        rows = [
            ('Method', "Accuracy"),
            ("Proposed PResNet", cfg.presnetacc),
            ("ResNet", cfg.exresnetacc),
            ("CNN", cfg.excnnacc),
            ("DNN", cfg.exdnnacc),
            ("ANN", cfg.exannacc),
        ]
        for row in rows:
            ws.append(row)
        data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=6)
        titles = Reference(ws, min_col=1, min_row=2, max_row=6)
        chart = BarChart3D()
        chart.title = "Accuracy"
        chart.add_data(data=data, titles_from_data=True)
        chart.set_categories(titles)
        chart.x_axis.title = "Classification algm"
        chart.y_axis.title = "Accuracy(%)"
        ws.add_chart(chart, "E5")
        wb.save("..\\Run\\Result\\accuracy.xlsx")
        print("(Accuracy)\n")
        x1 = PrettyTable.PrettyTable()
        x1.field_names = ['Method', "Accuracy"]
        x1.add_row(["Proposed PResNet", cfg.presnetacc])
        x1.add_row(["ResNet", cfg.exresnetacc])
        x1.add_row(["CNN", cfg.excnnacc])
        x1.add_row(["DNN", cfg.exdnnacc])
        x1.add_row(["ANN", cfg.exannacc])
        print(x1.get_string(title=""))

        fig = plt.figure(figsize=(8, 5))
        colors = ["green", "olive", "gray", "brown", "cyan"]
        X = ['Proposed PResNet', 'ResNet', 'CNN', 'DNN', 'ANN']
        Y = [85,88,90,95,98]
        # Accuracy = [round(int(cfg.presnetacc)), round(int(cfg.exresnetacc)), round(int(cfg.excnnacc)), round(int(cfg.exdnnacc)), round(int(cfg.exannacc))]
        Accuracy = [((cfg.presnetacc)), ((cfg.exresnetacc)), ((cfg.excnnacc)), ((cfg.exdnnacc)), ((cfg.exannacc))]
        X_axis = np.arange(len(X))
        plt.bar(X, Accuracy, color=colors)
        plt.xticks(X_axis, X, font="Times New Roman")
        plt.ylim(80,100)
        plt.yticks(font="Times New Roman")
        # plt.yticks()
        plt.xlabel("Techniques", font="Times New Roman", fontweight="bold")
        plt.ylabel("Percentage(%)", font="Times New Roman", fontweight="bold")
        plt.title("Accuracy", font="Times New Roman", fontweight="bold")
        plt.savefig("..\\Run\\Result\\Accuracy.png")
        plt.show()

    # precision()

    def precision(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        rows = [
            ('Method', "Precision"),
            ("Proposed  PResNet ", cfg.presnetpre),
            ("ResNet", cfg.exresnetpre),
            ("CNN", cfg.excnnpre),
            ("DNN", cfg.exdnnpre),
            ("ANN", cfg.exannpre),
        ]
        for row in rows:
            ws.append(row)
        data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=6)
        titles = Reference(ws, min_col=1, min_row=2, max_row=6)
        chart = BarChart3D()
        chart.title = "Precision"
        chart.add_data(data=data, titles_from_data=True)
        chart.set_categories(titles)
        chart.x_axis.title = "Classification algm"
        chart.y_axis.title = "Precision(%)"
        ws.add_chart(chart, "E5")
        wb.save("..\\Run\\Result\\precision.xlsx")
        print(
            "(Precision)\n")
        x1 = PrettyTable.PrettyTable()
        x1.field_names = ['Method', "Precision"]
        x1.add_row(["Proposed PResNet", cfg.presnetpre])
        x1.add_row(["ResNet", cfg.exresnetpre])
        x1.add_row(["CNN", cfg.excnnpre])
        x1.add_row(["DNN", cfg.exdnnpre])
        x1.add_row(["ANN", cfg.exannpre])
        print(x1.get_string(title=""))

        fig = plt.figure(figsize=(8, 5))
        colors = ["green", "orange", "blue", "maroon", "lightslategrey"]
        X = ['Proposed PResNet', 'ResNet', 'CNN', 'DNN', 'ANN']
        Precision = [(cfg.presnetpre), (cfg.exresnetpre), (cfg.excnnpre), (cfg.exdnnpre),(cfg.exannpre)]
        X_axis = np.arange(len(X))
        plt.bar(X, Precision, color=colors)
        plt.xticks(X_axis, X, font="Times New Roman")
        plt.ylim(80,100)
        plt.yticks(font="Times New Roman")
        plt.xlabel("Techniques", font="Times New Roman", fontweight="bold")
        plt.ylabel("Percentage(%)", font="Times New Roman", fontweight="bold")
        plt.title("Precision", font="Times New Roman", fontweight="bold")
        plt.savefig("..\\Run\\Result\\Precision.png")
        plt.show()

    # recall()

    def recall(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        rows = [
            ('Method', "Recall"),
            ("Proposed PResNet", cfg.presnetrecall),
            ("ResNet", cfg.exresnetrecall),
            ("CNN", cfg.excnnrecall),
            ("DNN", cfg.exdnnrecall),
            ("ANN", cfg.exannrecall),
        ]
        for row in rows:
            ws.append(row)
        data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=6)
        titles = Reference(ws, min_col=1, min_row=2, max_row=6)
        chart = BarChart3D()
        chart.title = "Recall"
        chart.add_data(data=data, titles_from_data=True)
        chart.set_categories(titles)
        chart.x_axis.title = "Classification algm"
        chart.y_axis.title = "Recall(%)"
        ws.add_chart(chart, "E5")
        wb.save("..\\Run\\Result\\recall.xlsx")
        print(
            "(Recall)\n")
        x1 = PrettyTable.PrettyTable()
        x1.field_names = ['Method', "Recall"]
        x1.add_row(["Proposed PResNet", cfg.presnetrecall])
        x1.add_row(["ResNet", cfg.exresnetrecall])
        x1.add_row(["CNN", cfg.excnnrecall])
        x1.add_row(["DNN", cfg.exdnnrecall])
        x1.add_row(["ANN", cfg.exannrecall])
        print(x1.get_string(title=""))

        fig = plt.figure(figsize=(8, 5))
        colors = ["blue", "brown", "olive", "orange", "cyan"]
        X = ['Proposed PResNet', 'ResNet', 'CNN', 'DNN', 'ANN']
        Recall = [(cfg.presnetrecall), (cfg.exresnetrecall), (cfg.excnnrecall), (cfg.exdnnrecall), (cfg.exannrecall)]
        X_axis = np.arange(len(X))
        plt.bar(X, Recall, color=colors)
        plt.xticks(X_axis, X, font="Times New Roman")
        plt.ylim(80,100)
        plt.yticks(font="Times New Roman")
        plt.xlabel("Techniques", font="Times New Roman", fontweight="bold")
        plt.ylabel("Percentage(%)", font="Times New Roman", fontweight="bold")
        plt.title("Recall", font="Times New Roman", fontweight="bold")
        plt.savefig("..\\Run\\Result\\Recall.png")
        plt.show()

    # f-measure()

    def f_measure(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        rows = [
            ('Method', "f-score"),
            ("Proposed PResNet ", cfg.presnetfscore),
            ("ResNet", cfg.exresnetfscore),
            ("CNN", cfg.excnnfscore),
            ("DNN", cfg.exdnnfscore),
            ("ANN", cfg.exannfscore),
        ]
        for row in rows:
            ws.append(row)
        data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=6)
        titles = Reference(ws, min_col=1, min_row=2, max_row=6)
        chart = BarChart3D()
        chart.title = "F-measure"
        chart.add_data(data=data, titles_from_data=True)
        chart.set_categories(titles)
        chart.x_axis.title = "Classification algm"
        chart.y_axis.title = "F-measure(%)"
        ws.add_chart(chart, "E5")
        wb.save("..\\Run\\Result\\F-measure.xlsx")
        print(
            "(F-Measure)\n")
        x1 = PrettyTable.PrettyTable()
        x1.field_names = ['Method', "F-measure"]
        x1.add_row(["Proposed PResNet", cfg.presnetfscore])
        x1.add_row(["ResNet", cfg.exresnetfscore])
        x1.add_row(["CNN", cfg.excnnfscore])
        x1.add_row(["DNN", cfg.exdnnfscore])
        x1.add_row(["ANN", cfg.exannfscore])
        print(x1.get_string(title=""))

        fig = plt.figure(figsize=(8, 5))
        colors = ["orangered", "darkgoldenrod", "mediumaquamarine", "slateblue", "purple"]
        X = ['Proposed PResNet', 'ResNet', 'CNN', 'DNN', 'ANN']
        Fscore = [(cfg.presnetfscore), (cfg.exresnetfscore), (cfg.excnnfscore), (cfg.exdnnfscore), (cfg.exannfscore)]
        X_axis = np.arange(len(X))
        plt.bar(X, Fscore, color=colors)
        plt.xticks(X_axis, X, font="Times New Roman")
        plt.ylim(80,100)
        plt.yticks(font="Times New Roman")
        plt.xlabel("Techniques", font="Times New Roman", fontweight="bold")
        plt.ylabel("Percentage(%)", font="Times New Roman", fontweight="bold")
        plt.title("F-score", font="Times New Roman", fontweight="bold")
        plt.savefig("..\\Run\\Result\\FScore.png")
        plt.show()

    # sensitivity()

    def sensitivity(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        rows = [
            ('Method', "sensitivity"),
            ("Proposed PResNet ",cfg.presnetsens),
            ("ResNet", cfg.exresnetsens),
            ("CNN", cfg.excnnsens),
            ("DNN", cfg.exdnnsens),
            ("ANN", cfg.exannsens),
        ]
        for row in rows:
            ws.append(row)
        data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=6)
        titles = Reference(ws, min_col=1, min_row=2, max_row=6)
        chart = BarChart3D()
        chart.title = "Sensitivity"
        chart.add_data(data=data, titles_from_data=True)
        chart.set_categories(titles)
        chart.x_axis.title = "Classification algm"
        chart.y_axis.title = "Sensitivity(%)"
        ws.add_chart(chart, "E5")
        wb.save("..\\Run\\Result\\Sensitivity.xlsx")
        print(
            "Sensitivity")
        x1 = PrettyTable.PrettyTable()
        x1.field_names = ['Method', "Sensitivity"]
        x1.add_row(["Proposed PResNet", cfg.presnetsens])
        x1.add_row(["ResNet", cfg.exresnetsens])
        x1.add_row(["CNN", cfg.excnnsens])
        x1.add_row(["DNN", cfg.exdnnsens])
        x1.add_row(["ANN", cfg.exannsens])
        print(x1.get_string(title=""))

        fig = plt.figure(figsize=(8, 5))
        X = ['Proposed PResNet', 'ResNet', 'CNN', 'DNN', 'ANN']
        Sensitivity = [(cfg.presnetsens),(cfg.exresnetsens), (cfg.excnnsens), (cfg.exdnnsens), (cfg.exannsens)]
        X_axis = np.arange(len(X))
        plt.plot(X, Sensitivity, color="darkcyan", marker = "*")
        plt.xticks(X_axis, X, font="Times New Roman")
        plt.ylim(80,100)
        plt.yticks(font="Times New Roman")
        plt.xlabel("Techniques", font="Times New Roman", fontweight="bold")
        plt.ylabel("Percentage(%)", font="Times New Roman", fontweight="bold")
        plt.title("Sensitivity", font="Times New Roman", fontweight="bold")
        plt.savefig("..\\Run\\Result\\Sensitivity.png")
        plt.show()

    # specificity()

    def specificity(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        rows = [
            ('Method', "specificity"),
            ("Proposed PResNet ", cfg.presnetspec),
            ("ResNet", cfg.exresnetspec),
            ("CNN", cfg.excnnspec),
            ("DNN", cfg.exdnnspec),
            ("ANN", cfg.exannspec),
        ]
        for row in rows:
            ws.append(row)
        data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=6)
        titles = Reference(ws, min_col=1, min_row=2, max_row=6)
        chart = BarChart3D()
        chart.title = "Specificity"
        chart.add_data(data=data, titles_from_data=True)
        chart.set_categories(titles)
        chart.x_axis.title = "Classification algm"
        chart.y_axis.title = "Specificity(%)"
        ws.add_chart(chart, "E5")
        wb.save("..\\Run\\Result\\Specificity.xlsx")
        print(
            "(Specificity)\n")
        x1 = PrettyTable.PrettyTable()
        x1.field_names = ['Method', "Specificity"]
        x1.add_row(["Proposed PResNet", cfg.presnetspec])
        x1.add_row(["ResNet", cfg.exresnetspec])
        x1.add_row(["CNN", cfg.excnnspec])
        x1.add_row(["DNN", cfg.exdnnspec])
        x1.add_row(["ANN", cfg.exannspec])
        print(x1.get_string(title=""))

        fig = plt.figure(figsize=(8, 5))
        colors = ["green", "orange", "blue", "maroon", "cyan"]
        X = ['Proposed PResNet', 'ResNet', 'CNN', 'DNN', 'ANN']
        Specificity = [(cfg.presnetspec), (cfg.exresnetspec), (cfg.excnnspec), (cfg.exdnnspec), (cfg.exannspec)]
        X_axis = np.arange(len(X))
        plt.bar(X, Specificity, color=colors)
        plt.xticks(X_axis, X, font="Times New Roman")
        plt.ylim(80,100)
        plt.yticks(font="Times New Roman")
        plt.xlabel("Techniques", font="Times New Roman", fontweight="bold")
        plt.ylabel("Percentage(%)", font="Times New Roman", fontweight="bold")
        plt.title("Specificity", font="Times New Roman", fontweight="bold")
        plt.savefig("..\\Run\\Result\\Specificity.png")
        plt.show()

    # Training time

    # def training_time(self):
    #     wb = openpyxl.Workbook()
    #     ws = wb.active
    #     rows = [
    #         ('Method', "Training_time"),
    #         ("Proposed PResNet ", cfg.presnet_trtime),
    #         ("ResNet", cfg.exresnet_trtime),
    #         ("CNN", cfg.excnntrtime),
    #         ("DNN", cfg.exdnntrtime),
    #         ("ANN", cfg.exanntrtime),
    #     ]
    #     for row in rows:
    #         ws.append(row)
    #     data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=6)
    #     titles = Reference(ws, min_col=1, min_row=2, max_row=6)
    #     chart = BarChart3D()
    #     chart.title = "Training time (ms)"
    #     chart.add_data(data=data, titles_from_data=True)
    #     chart.set_categories(titles)
    #     chart.x_axis.title = "Classification algm"
    #     chart.y_axis.title = "Training time(%)"
    #     ws.add_chart(chart, "E5")
    #     if not os.path.exists("..\\Result\\Training_time.xlsx"):
    #         wb.save("..\\Run\\Result\\Training_time.xlsx")
    #     print(
    #         "(Training time)\n")
    #     x1 = PrettyTable.PrettyTable()
    #     x1.field_names = ['Method', "Training time"]
    #     x1.add_row(["Proposed PResNet", cfg.presnet_trtime])
    #     x1.add_row(["ResNet", cfg.exresnet_trtime])
    #     x1.add_row(["CNN", cfg.excnntrtime])
    #     x1.add_row(["DNN", cfg.exdnntrtime])
    #     x1.add_row(["ANN", cfg.exanntrtime])
    #     print(x1.get_string(title=""))
    #
    #     fig = plt.figure(figsize=(8, 5))
    #     X = ['Proposed PResNet', 'ResNet', 'CNN', 'DNN', 'ANN']
    #     time = [(cfg.presnet_trtime), (cfg.exresnet_trtime), (cfg.excnntrtime), (cfg.exdnntrtime), (cfg.exanntrtime)]
    #     X_axis = np.arange(len(X))
    #     plt.plot(X, time, color="dodgerblue", marker="o")
    #     plt.xticks(X_axis, X, font="Times New Roman")
    #     # plt.ylim(80,100)
    #     plt.yticks(font="Times New Roman")
    #     plt.xlabel("Techniques", font="Times New Roman", fontweight="bold")
    #     plt.ylabel("time(ms)", font="Times New Roman", fontweight="bold")
    #     plt.title("TrainingTime", font="Times New Roman", fontweight="bold")
    #     if not os.path.exists("..\\Result\\Training_time.png"):
    #         plt.savefig("..\\Run\\Result\\Training_time.png")
    #     plt.show()

    # FPR

    def Fpr(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        rows = [
            ('Method', "FPR"),
            ("Proposed PResNet ", cfg.presnetfpr),
            ("ResNet", cfg.exresnetfpr),
            ("CNN", cfg.excnnfpr),
            ("DNN", cfg.exdnnfpr),
            ("ANN", cfg.exannfpr),
        ]
        for row in rows:
            ws.append(row)
        data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=6)
        titles = Reference(ws, min_col=1, min_row=2, max_row=6)
        chart = BarChart3D()
        chart.title = "FPR"
        chart.add_data(data=data, titles_from_data=True)
        chart.set_categories(titles)
        chart.x_axis.title = "Classification algm"
        chart.y_axis.title = "FPR"
        ws.add_chart(chart, "E5")
        wb.save("..\\Run\\Result\\FPR.xlsx")
        print(
            "(FPR)\n")
        x1 = PrettyTable.PrettyTable()
        x1.field_names = ['Method', "FPR"]
        x1.add_row(["Proposed PResNet", cfg.presnetfpr])
        x1.add_row(["ResNet", cfg.exresnetfpr])
        x1.add_row(["CNN", cfg.excnnfpr])
        x1.add_row(["DNN", cfg.exdnnfpr])
        x1.add_row(["ANN", cfg.exannfpr])
        print(x1.get_string(title=""))

        fig = plt.figure(figsize=(8, 5))
        X = ['Proposed PResNet', 'ResNet', 'CNN', 'DNN', 'ANN']
        time = [(cfg.presnetfpr), (cfg.exresnetfpr), (cfg.excnnfpr), (cfg.exdnnfpr), (cfg.exannfpr)]
        X_axis = np.arange(len(X))
        plt.plot(X, time, color="lightseagreen", marker="o")
        plt.xticks(X_axis, X, font="Times New Roman")
        plt.ylim(0,0.1)
        plt.yticks(font="Times New Roman")
        plt.xlabel("Techniques", font="Times New Roman", fontweight="bold")
        plt.ylabel("FPR", font="Times New Roman", fontweight="bold")
        plt.title("FPR", font="Times New Roman", fontweight="bold")
        plt.savefig("..\\Run\\Result\\FPR.png")
        plt.show()

    #FNR

    def Fnr(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        rows = [
            ('Method', "FNR"),
            ("Proposed PResNet ", cfg.presnetfnr),
            ("ResNet", cfg.exresnetfnr),
            ("CNN", cfg.excnnfnr),
            ("DNN", cfg.exdnnfnr),
            ("ANN", cfg.exannfnr),
        ]
        for row in rows:
            ws.append(row)
        data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=6)
        titles = Reference(ws, min_col=1, min_row=2, max_row=6)
        chart = BarChart3D()
        chart.title = "FNR"
        chart.add_data(data=data, titles_from_data=True)
        chart.set_categories(titles)
        chart.x_axis.title = "Classification algm"
        chart.y_axis.title = "FNR"
        ws.add_chart(chart, "E5")
        wb.save("..\\Run\\Result\\FNR.xlsx")
        print(
            "(FNR)\n")
        x1 = PrettyTable.PrettyTable()
        x1.field_names = ['Method', "FNR"]
        x1.add_row(["Proposed PResNet", cfg.presnetfnr])
        x1.add_row(["ResNet", cfg.exresnetfnr])
        x1.add_row(["CNN", cfg.excnnfnr])
        x1.add_row(["DNN", cfg.exdnnfnr])
        x1.add_row(["ANN", cfg.exannfnr])
        print(x1.get_string(title=""))

        fig = plt.figure(figsize=(8, 5))
        colors = ["darkred", "coral", "saddlebrown", "olivedrab", "indigo"]
        X = ['Proposed PResNet', 'ResNet', 'CNN', 'DNN', 'ANN']
        Specificity = [(cfg.presnetfnr), (cfg.exresnetfnr), (cfg.excnnfnr), (cfg.exdnnfnr), (cfg.exannfnr)]
        X_axis = np.arange(len(X))
        plt.bar(X, Specificity, color=colors)
        plt.xticks(X_axis, X, font="Times New Roman")
        plt.ylim(0,0.2)
        plt.yticks(font="Times New Roman")
        plt.xlabel("Techniques", font="Times New Roman", fontweight="bold")
        plt.ylabel("FNR", font="Times New Roman", fontweight="bold")
        plt.title("FNR", font="Times New Roman", fontweight="bold")
        plt.savefig("..\\Run\\Result\\FNR.png")
        plt.show()

    #FRR

    def Frr(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        rows = [
            ('Method', "FRR"),
            ("Proposed PResNet ", cfg.presnetfrr),
            ("ResNet", cfg.exresnetfrr),
            ("CNN", cfg.excnnfrr),
            ("DNN", cfg.exdnnfrr),
            ("ANN", cfg.exannfrr),
        ]
        for row in rows:
            ws.append(row)
        data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=6)
        titles = Reference(ws, min_col=1, min_row=2, max_row=6)
        chart = BarChart3D()
        chart.title = "FRR"
        chart.add_data(data=data, titles_from_data=True)
        chart.set_categories(titles)
        chart.x_axis.title = "Classification algm"
        chart.y_axis.title = "FRR"
        ws.add_chart(chart, "E5")
        wb.save("..\\Run\\Result\\FRR.xlsx")
        print(
            "(FRR)\n")
        x1 = PrettyTable.PrettyTable()
        x1.field_names = ['Method', "FRR"]
        x1.add_row(["Proposed PResNet", cfg.presnetfrr])
        x1.add_row(["ResNet", cfg.exresnetfrr])
        x1.add_row(["CNN", cfg.excnnfrr])
        x1.add_row(["DNN", cfg.exdnnfrr])
        x1.add_row(["ANN", cfg.exannfrr])
        print(x1.get_string(title=""))

        fig = plt.figure(figsize=(8, 5))
        X = ['Proposed PResNet', 'ResNet', 'CNN', 'DNN', 'ANN']
        time = [(cfg.presnetfrr), (cfg.exresnetfrr), (cfg.excnnfrr), (cfg.exdnnfrr), (cfg.exannfrr)]
        X_axis = np.arange(len(X))
        plt.plot(X, time, color="violet", marker="o")
        plt.xticks(X_axis, X, font="Times New Roman")
        plt.ylim(0,0.1)
        plt.yticks(font="Times New Roman")
        plt.xlabel("Techniques", font="Times New Roman", fontweight="bold")
        plt.ylabel("FRR", font="Times New Roman", fontweight="bold")
        plt.title("FRR", font="Times New Roman", fontweight="bold")
        plt.savefig("..\\Run\\Result\\FRR.png")
        plt.show()

    #error rate

    def Error_rate(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        rows = [
            ('Method', "Error_rate"),
            ("Proposed PResNet ", cfg.presneterrorrate),
            ("ResNet", cfg.exresneterrorrate),
            ("CNN", cfg.excnnerrorrate),
            ("DNN", cfg.exdnnerrorrate),
            ("ANN", cfg.exannerrorrate),
        ]
        for row in rows:
            ws.append(row)
        data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=6)
        titles = Reference(ws, min_col=1, min_row=2, max_row=6)
        chart = BarChart3D()
        chart.title = "Error Rate"
        chart.add_data(data=data, titles_from_data=True)
        chart.set_categories(titles)
        chart.x_axis.title = "Classification algm"
        chart.y_axis.title = "Error rate"
        ws.add_chart(chart, "E5")
        wb.save("..\\Run\\Result\\ErrorRate.xlsx")
        print(
            "(Error rate)\n")
        x1 = PrettyTable.PrettyTable()
        x1.field_names = ['Method', "Error Rate"]
        x1.add_row(["Proposed PResNet", cfg.presneterrorrate])
        x1.add_row(["ResNet", cfg.exresneterrorrate])
        x1.add_row(["CNN", cfg.excnnerrorrate])
        x1.add_row(["DNN", cfg.exdnnerrorrate])
        x1.add_row(["ANN", cfg.exannerrorrate])
        print(x1.get_string(title=""))

        fig = plt.figure(figsize=(8, 5))
        X = ['Proposed PResNet', 'ResNet', 'CNN', 'DNN', 'ANN']
        time = [(cfg.presneterrorrate), (cfg.exresneterrorrate), (cfg.excnnerrorrate), (cfg.exdnnerrorrate), (cfg.exannerrorrate)]
        X_axis = np.arange(len(X))
        plt.plot(X, time, color="darkslategrey", marker="*")
        plt.xticks(X_axis, X, font="Times New Roman")
        plt.ylim(0,0.2)
        plt.yticks(font="Times New Roman")
        plt.xlabel("Techniques", font="Times New Roman", fontweight="bold")
        plt.ylabel("Error rate", font="Times New Roman", fontweight="bold")
        plt.title("Error Rate", font="Times New Roman", fontweight="bold")
        plt.savefig("..\\Run\\Result\\ErrorRate.png")
        plt.show()

    #PSNR
    def PSNR(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        rows = [
            ('Method', "PSNR"),
            ("Proposed I-ADF", cfg.iadfpsnr),
            ("ADF", cfg.adfpsnr),
            ("GF", cfg.gfpsnr),
            ("MF", cfg.mfpsnr),
            ("BF", cfg.bfpsnr),
        ]
        for row in rows:
            ws.append(row)
        data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=6)
        titles = Reference(ws, min_col=1, min_row=2, max_row=6)
        chart = BarChart3D()
        chart.title = "PSNR"
        chart.add_data(data=data, titles_from_data=True)
        chart.set_categories(titles)
        chart.x_axis.title = "Denoising techniques"
        chart.y_axis.title = "PSNR"
        ws.add_chart(chart, "E5")
        wb.save("..\\Run\\Result\\PSNR.xlsx")
        print(
            "(PSNR)\n")
        x1 = PrettyTable.PrettyTable()
        x1.field_names = ['Method', "PSNR"]
        x1.add_row(["Proposed I-ADF", cfg.iadfpsnr])
        x1.add_row(["ADF", cfg.adfpsnr])
        x1.add_row(["GF", cfg.gfpsnr])
        x1.add_row(["MF", cfg.mfpsnr])
        x1.add_row(["BF", cfg.bfpsnr])
        print(x1.get_string(title=""))

        fig = plt.figure(figsize=(8, 5))
        X = ['Proposed I-ADF', 'ADF', 'GF', 'MF', 'BF']
        PSNR = [(cfg.iadfpsnr), (cfg.adfpsnr), (cfg.gfpsnr), (cfg.mfpsnr), (cfg.bfpsnr)]
        X_axis = np.arange(len(X))
        plt.plot(X, PSNR, color="darkslategrey", marker="*")
        plt.xticks(X_axis, X, font="Times New Roman")
        plt.ylim(0,35)
        plt.yticks(font="Times New Roman")
        plt.xlabel("Techniques", font="Times New Roman", fontweight="bold")
        plt.ylabel("Values", font="Times New Roman", fontweight="bold")
        plt.title("PSNR", font="Times New Roman", fontweight="bold")
        plt.savefig("..\\Run\\Result\\PSNR.png")
        plt.show()

    #MSE
    def MSE(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        rows = [
            ('Method', "MSE"),
            ("Proposed I-ADF ", cfg.iadfmse),
            ("ADF", cfg.adfmse),
            ("GF", cfg.gfmse),
            ("MF", cfg.mfmse),
            ("BF", cfg.bfmse),
        ]
        for row in rows:
            ws.append(row)
        data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=6)
        titles = Reference(ws, min_col=1, min_row=2, max_row=6)
        chart = BarChart3D()
        chart.title = "Specificity"
        chart.add_data(data=data, titles_from_data=True)
        chart.set_categories(titles)
        chart.x_axis.title = "Denoising techniques"
        chart.y_axis.title = "MSE"
        ws.add_chart(chart, "E5")
        wb.save("..\\Run\\Result\\MSE.xlsx")
        print(
            "(MSE)\n")
        x1 = PrettyTable.PrettyTable()
        x1.field_names = ['Method', "MSE"]
        x1.add_row(["Proposed I-ADF", cfg.iadfmse])
        x1.add_row(["ADF", cfg.adfmse])
        x1.add_row(["GF", cfg.gfmse])
        x1.add_row(["MF", cfg.mfmse])
        x1.add_row(["BF", cfg.bfmse])
        print(x1.get_string(title=""))

        fig = plt.figure(figsize=(8, 5))
        colors = ["green", "orange", "blue", "maroon", "cyan"]
        X = ['Proposed I-ADF', 'ADF', 'GF', 'MF', 'BF']
        MSE = [(cfg.iadfmse), (cfg.adfmse), (cfg.gfmse), (cfg.mfmse), (cfg.bfmse)]
        X_axis = np.arange(len(X))
        plt.bar(X, MSE, color=colors)
        plt.xticks(X_axis, X, font="Times New Roman")
        plt.ylim(0,10)
        plt.yticks(font="Times New Roman")
        plt.xlabel("Techniques", font="Times New Roman", fontweight="bold")
        plt.ylabel("Error values", font="Times New Roman", fontweight="bold")
        plt.title("MSE", font="Times New Roman", fontweight="bold")
        plt.savefig("..\\Run\\Result\\MSE.png")
        plt.show()

    #SSIM

    def SSIM(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        rows = [
            ('Method', "SSIM"),
            ("Proposed I-ADF ", cfg.iadfssim),
            ("ADF", cfg.adfssim),
            ("GF", cfg.gfssim),
            ("MF", cfg.mfssim),
            ("BF", cfg.bfssim),
        ]
        for row in rows:
            ws.append(row)
        data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=6)
        titles = Reference(ws, min_col=1, min_row=2, max_row=6)
        chart = BarChart3D()
        chart.title = "SSIM"
        chart.add_data(data=data, titles_from_data=True)
        chart.set_categories(titles)
        chart.x_axis.title = "Denoising techniques"
        chart.y_axis.title = "SSIM"
        ws.add_chart(chart, "E5")
        wb.save("..\\Run\\Result\\SSIM.xlsx")
        print(
            "(SSIM)\n")
        x1 = PrettyTable.PrettyTable()
        x1.field_names = ['Method', "SSIM"]
        x1.add_row(["Proposed I-ADF", cfg.iadfssim])
        x1.add_row(["ADF", cfg.adfssim])
        x1.add_row(["GF", cfg.gfssim])
        x1.add_row(["MF", cfg.mfssim])
        x1.add_row(["BF", cfg.bfssim])
        print(x1.get_string(title=""))

        fig = plt.figure(figsize=(8, 5))
        colors = ["green", "orange", "blue", "maroon", "cyan"]
        X = ['Proposed I-ADF', 'ADF', 'GF', 'MF', 'BF']
        SSIM = [(cfg.iadfssim), (cfg.adfssim), (cfg.gfssim), (cfg.mfssim), (cfg.bfssim)]
        X_axis = np.arange(len(X))
        plt.bar(X, SSIM, color=colors)
        plt.xticks(X_axis, X, font="Times New Roman")
        plt.ylim(0,1)
        plt.yticks(font="Times New Roman")
        plt.xlabel("Techniques", font="Times New Roman", fontweight="bold")
        plt.ylabel("values", font="Times New Roman", fontweight="bold")
        plt.title("SSIM", font="Times New Roman", fontweight="bold")
        plt.savefig("..\\Run\\Result\\SSIM.png")
        plt.show()

    def dice_score(self):

        wb = openpyxl.Workbook()
        ws = wb.active
        rows = [
            ('Method', "Dice Score"),
            ("Proposed BRGS", cfg.brgsds),
            ("RGS", cfg.rgsds),
            ("WS", cfg.wsds),
            ("OS", cfg.osds),
            ("KMA", cfg.kmads),
        ]
        for row in rows:
            ws.append(row)
        data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=6)
        titles = Reference(ws, min_col=1, min_row=2, max_row=6)
        chart = BarChart3D()
        chart.title = "Dice Score"
        chart.add_data(data=data, titles_from_data=True)
        chart.set_categories(titles)
        chart.x_axis.title = "Segmentation techniques"
        chart.y_axis.title = "Dice Score"
        ws.add_chart(chart, "E5")
        wb.save("..\\Run\\Result\\Dice_Score.xlsx")
        print("(Dice Score)\n")
        x1 = PrettyTable.PrettyTable()
        x1.field_names = ['Method', "Dice Score"]
        x1.add_row(["Proposed BRGS", cfg.brgsds])
        x1.add_row(["RGS", cfg.rgsds])
        x1.add_row(["WS", cfg.wsds])
        x1.add_row(["OS", cfg.osds])
        x1.add_row(["KMA", cfg.kmads])
        print(x1.get_string(title=""))

        fig = plt.figure(figsize=(8, 5))
        colors = ["green", "olive", "gray", "brown", "cyan"]
        X = ['Proposed BRGS', 'RGS', 'WS', 'OS', 'KMA']
        Accuracy = [((cfg.brgsds)), ((cfg.rgsds)), ((cfg.wsds)), ((cfg.osds)), ((cfg.kmads))]
        X_axis = np.arange(len(X))
        plt.bar(X, Accuracy, color=colors)
        plt.xticks(X_axis, X, font="Times New Roman")
        plt.ylim(0,1)
        plt.yticks(font="Times New Roman")
        plt.xlabel("Techniques", font="Times New Roman", fontweight="bold")
        plt.ylabel("Values", font="Times New Roman", fontweight="bold")
        plt.title("Dice Score", font="Times New Roman", fontweight="bold")
        plt.savefig("..\\Run\\Result\\Dice_score.png")
        plt.show()



